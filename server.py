import http.server
import json
import os
import threading
import hashlib
import secrets
import io
import csv
from datetime import datetime, timedelta
from collections import defaultdict
from urllib.request import Request, urlopen
from urllib.parse import urlparse, parse_qs

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# GitHub Gist cache
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "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")
GIST_DESCRIPTION = "sanpretta-dashboard-cache"
_gist_id = None  # cached after first lookup

def _gist_request(method, url, data=None):
    body = json.dumps(data).encode() if data else None
    req = Request(url, data=body, method=method, headers={
        "Authorization": f"token {GITHUB_TOKEN}",
        "Accept": "application/vnd.github+json",
        "Content-Type": "application/json",
        "User-Agent": "SanPretta-Dashboard"
    })
    with urlopen(req) as resp:
        return json.loads(resp.read())

def _find_gist_id():
    global _gist_id
    if _gist_id:
        return _gist_id
    try:
        gists = _gist_request("GET", "https://api.github.com/gists?per_page=100")
        for g in gists:
            if g.get("description") == GIST_DESCRIPTION:
                _gist_id = g["id"]
                print(f"  [Gist] Encontrado gist existente: {_gist_id}")
                return _gist_id
    except Exception as e:
        print(f"  [Gist] Error buscando gist: {e}")
    return None

def drive_load_cache():
    """Load summary cache + raw orders from GitHub Gist."""
    if not GITHUB_TOKEN:
        return None
    try:
        gist_id = _find_gist_id()
        if not gist_id:
            print("  [Gist] Sin caché previo")
            return None
        gist = _gist_request("GET", f"https://api.github.com/gists/{gist_id}")

        # Load summary
        raw_url = gist["files"]["cache.json"]["raw_url"]
        req = Request(raw_url, headers={"Authorization": f"token {GITHUB_TOKEN}", "User-Agent": "SanPretta-Dashboard"})
        with urlopen(req) as resp:
            data = json.loads(resp.read())
        print(f"  [Gist] Caché cargado OK ({len(data.get('products', []))} productos)")

        # Load raw orders if available (avoids re-fetching from Tiendanube on startup)
        if "raw_orders.json" in gist["files"]:
            try:
                raw_url2 = gist["files"]["raw_orders.json"]["raw_url"]
                req2 = Request(raw_url2, headers={"Authorization": f"token {GITHUB_TOKEN}", "User-Agent": "SanPretta-Dashboard"})
                with urlopen(req2) as resp2:
                    raw_data = json.loads(resp2.read())
                _raw_cache["all_orders"] = raw_data.get("all_orders", [])
                _raw_cache["last_updated"] = raw_data.get("last_updated")
                # Siempre refrescar productos desde la API para tener stock actualizado
                print("  [Gist] Actualizando productos desde Tiendanube (stock en tiempo real)...")
                fresh_products = fetch_products()
                _raw_cache["products"] = fresh_products
                _raw_cache["variant_map"], _raw_cache["product_names"] = build_variant_map(fresh_products)
                print(f"  [Gist] Órdenes crudas: {len(_raw_cache['all_orders'])} · Productos actualizados: {len(fresh_products)}")
            except Exception as e:
                print(f"  [Gist] No se pudieron cargar órdenes crudas: {e}")

        return data
    except Exception as e:
        print(f"  [Gist] Error cargando caché: {e}")
        return None

def drive_save_cache(data):
    """Save summary cache + raw orders to GitHub Gist."""
    global _gist_id
    if not GITHUB_TOKEN:
        return
    try:
        summary_content = json.dumps(data, ensure_ascii=False)
        # Also save raw orders if available (for fast startup)
        raw_content = None
        if _raw_cache["all_orders"] is not None:
            raw_data = {
                "all_orders": _raw_cache["all_orders"],
                "products": _raw_cache["products"],
                "last_updated": _raw_cache["last_updated"]
            }
            raw_content = json.dumps(raw_data, ensure_ascii=False)
            print(f"  [Gist] Guardando {len(_raw_cache['all_orders'])} órdenes crudas")

        files = {"cache.json": {"content": summary_content}}
        if raw_content:
            files["raw_orders.json"] = {"content": raw_content}

        gist_id = _find_gist_id()
        if gist_id:
            _gist_request("PATCH", f"https://api.github.com/gists/{gist_id}", {"files": files})
            print("  [Gist] Caché actualizado en GitHub")
        else:
            result = _gist_request("POST", "https://api.github.com/gists", {
                "description": GIST_DESCRIPTION,
                "public": False,
                "files": files
            })
            _gist_id = result["id"]
            print(f"  [Gist] Caché creado en GitHub: {_gist_id}")
    except Exception as e:
        print(f"  [Gist] Error guardando caché: {e}")


STORE_ID = "87884"
TOKEN = "c8e480ded10aed4f9e1dd31fb15f8ba658c3b72b"
BASE_URL = f"https://api.tiendanube.com/v1/{STORE_ID}"
import os
PORT = int(os.environ.get("PORT", 8765))

# Auth
USERNAME = "juanman"
PASSWORD_HASH = hashlib.sha256("Emilia27Aurora07!".encode()).hexdigest()
SESSIONS = set()  # active session tokens

def check_session(handler):
    cookie = handler.headers.get("Cookie", "")
    for part in cookie.split(";"):
        part = part.strip()
        if part.startswith("sp_session="):
            token = part[len("sp_session="):]
            if token in SESSIONS:
                return True
    return False

LOGIN_HTML = '''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>San Pretta · Acceso</title>
<link href="https://fonts.googleapis.com/css2?family=Cormorant+Garamond:wght@400;500&family=Jost:wght@300;400&display=swap" rel="stylesheet">
<style>
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: "Jost", sans-serif; background: #f5ede6; min-height: 100vh; display: flex; align-items: center; justify-content: center; }
.card { background: #fdf8f5; border: 1px solid rgba(156,123,110,0.18); border-radius: 14px; padding: 2.5rem 2rem; width: 100%; max-width: 360px; }
.title { font-family: "Cormorant Garamond", serif; font-size: 28px; font-weight: 500; color: #3a2a24; text-align: center; margin-bottom: 0.25rem; }
.subtitle { font-size: 11px; text-transform: uppercase; letter-spacing: 0.15em; color: #b09080; text-align: center; margin-bottom: 2rem; }
label { font-size: 10px; text-transform: uppercase; letter-spacing: 0.1em; color: #b09080; display: block; margin-bottom: 4px; margin-top: 1rem; }
input { width: 100%; font-family: "Jost", sans-serif; font-size: 13px; padding: 9px 12px; border: 1px solid rgba(156,123,110,0.32); border-radius: 8px; background: #f5ede6; color: #3a2a24; outline: none; }
input:focus { border-color: #9c7b6e; }
button { width: 100%; margin-top: 1.5rem; font-family: "Jost", sans-serif; font-size: 12px; font-weight: 500; letter-spacing: 0.1em; text-transform: uppercase; color: #fdf8f5; background: #7a5c50; border: none; border-radius: 8px; padding: 11px; cursor: pointer; }
button:hover { background: #9c7b6e; }
.error { font-size: 12px; color: #a03030; text-align: center; margin-top: 1rem; background: #faeaea; padding: 8px; border-radius: 6px; }
</style>
</head>
<body>
<div class="card">
  <div class="title">San Pretta</div>
  <div class="subtitle">Dashboard de ventas</div>
  <form method="POST" action="/login">
    <label>Usuario</label>
    <input type="text" name="username" autocomplete="username" required>
    <label>Contraseña</label>
    <input type="password" name="password" autocomplete="current-password" required>
    <button type="submit">Ingresar</button>
    {error}
  </form>
</div>
</body>
</html>'''

HEADERS = {
    "Authentication": f"bearer {TOKEN}",
    "User-Agent": "DashboardSanPretta (dashboard@sanpretta.com)",
    "Content-Type": "application/json"
}

# Progress tracking
_progress = {"pct": 0, "msg": ""}

def set_progress(pct, msg):
    _progress["pct"] = pct
    _progress["msg"] = msg
    print(f"  [{pct}%] {msg}")

def api_get(url):
    req = Request(url, headers=HEADERS)
    with urlopen(req) as resp:
        return json.loads(resp.read())

def fetch_orders(days=None, date_from=None, date_to=None, progress_range=(0,50), label=""):
    results = []
    page = 1
    base = f"{BASE_URL}/orders?payment_status=paid&per_page=200"
    if date_from:
        base += f"&created_at_min={date_from}"
        if date_to:
            base += f"&created_at_max={date_to}"
    elif days:
        since = (datetime.now() - timedelta(days=days)).strftime("%Y-%m-%d")
        base += f"&created_at_min={since}"
    else:
        since = (datetime.now() - timedelta(days=730)).strftime("%Y-%m-%d")
        base += f"&created_at_min={since}"
    p_start, p_end = progress_range
    while True:
        data = api_get(f"{base}&page={page}")
        if not isinstance(data, list) or not data:
            break
        results.extend(data)
        # Estimate progress — we don't know total, so use pages as proxy (cap at 80% of range)
        est = min(p_start + int((page / max(page+2, 5)) * (p_end - p_start) * 0.9), p_end - 2)
        set_progress(est, f"{label}: {len(results)} pedidos cargados")
        if len(data) < 200:
            break
        page += 1
    set_progress(p_end, f"{label}: {len(results)} pedidos")
    return results

def fetch_products():
    results = []
    page = 1
    set_progress(2, "Cargando productos...")
    while True:
        data = api_get(f"{BASE_URL}/products?per_page=200&page={page}")
        if not isinstance(data, list) or not data:
            break
        results.extend(data)
        set_progress(4, f"Productos: {len(results)} cargados")
        if len(data) < 200:
            break
        page += 1
    set_progress(8, f"Productos: {len(results)} total")
    return results

def get_name(val):
    if isinstance(val, dict):
        return val.get("es") or val.get("pt") or next(iter(val.values()), "")
    return str(val or "")

def days_since(date_str):
    if not date_str:
        return 9999
    try:
        dt = datetime.fromisoformat(date_str.replace("Z", "+00:00")).replace(tzinfo=None)
        return (datetime.now() - dt).days
    except:
        return 9999

def safe_float(val):
    try:
        v = float(val or 0)
        return v if v > 0 else 0.0
    except:
        return 0.0

def build_variant_map(products):
    variant_map = {}
    product_names = {}
    for p in products:
        pid = p["id"]
        pname = get_name(p.get("name", ""))
        product_names[pid] = pname
        product_created = p.get("created_at", "")
        p_price = safe_float(p.get("price"))
        p_promo = safe_float(p.get("promotional_price"))
        if p_promo >= p_price: p_promo = 0.0
        images = p.get("images", [])
        product_image = images[0].get("src", "") if images and isinstance(images[0], dict) else ""
        for v in p.get("variants", []):
            vid = v["id"]
            parts = [get_name(val) if isinstance(val, dict) else str(val) for val in v.get("values", [])]
            vname = " / ".join(parts) if parts else ""
            try: stock = int(v.get("stock", 0) or 0)
            except: stock = 0
            v_created = v.get("created_at", "") or product_created
            # Use updated_at when stock > 0: reflects when stock was last added
            # This avoids marking recently restocked items as stagnant
            v_updated = v.get("updated_at", "") or v_created
            v_ref_date = v_created
            v_price = safe_float(v.get("price")) or p_price
            v_promo = safe_float(v.get("promotional_price")) or p_promo
            if v_promo >= v_price: v_promo = 0.0
            variant_map[vid] = {
                "product_id": pid,
                "product_name": pname,
                "variant_name": vname,
                "stock": stock,
                "days_in_catalog": days_since(v_ref_date),
                "image": product_image,
                "price": v_price,
                "promo_price": v_promo
            }
    return variant_map, product_names

def get_variants_with_sales(orders):
    """Return set of variant_ids that have ANY sales in the given orders."""
    result = set()
    for order in orders:
        for item in order.get("products", []):
            vid = item.get("variant_id")
            if vid and int(item.get("quantity", 0)) > 0:
                result.add(vid)
    return result

def _calc_historical_rate(dates):
    """Calculate real velocity from first sale to last sale (or today if recent)."""
    if not dates:
        return 0.0
    sorted_dates = sorted(dates)
    try:
        first = datetime.strptime(sorted_dates[0], "%Y-%m-%d")
        last = datetime.now()
        span = max((last - first).days, 1)
        return round(len(dates) / span, 3)
    except:
        return 0.0

# Raw data cache — fetched once, reused for any period
_raw_cache = {"products": None, "all_orders": None, "variant_map": None, "product_names": None, "last_updated": None}

def fetch_raw_data(incremental=False):
    """Fetch products and order history from API. If incremental=True and we have
    existing data, only fetch orders newer than last_updated and merge."""
    now_str = datetime.now().strftime("%Y-%m-%d")

    if incremental and _raw_cache["all_orders"] is not None and _raw_cache["last_updated"]:
        # Only fetch orders since last update
        last = _raw_cache["last_updated"]
        set_progress(5, f"Actualizando desde {last}...")
        products = fetch_products()
        variant_map, product_names = build_variant_map(products)
        new_orders = fetch_orders(date_from=last, date_to=now_str, progress_range=(10, 70), label="Nuevos pedidos")
        if new_orders:
            # Merge: add new orders, remove duplicates by order id
            existing_ids = {o["id"] for o in _raw_cache["all_orders"]}
            added = [o for o in new_orders if o["id"] not in existing_ids]
            _raw_cache["all_orders"] = _raw_cache["all_orders"] + added
            print(f"  [Cache] Incremental: +{len(added)} pedidos nuevos")
        else:
            print("  [Cache] Incremental: sin pedidos nuevos")
        _raw_cache["products"] = products
        _raw_cache["variant_map"] = variant_map
        _raw_cache["product_names"] = product_names
        _raw_cache["last_updated"] = now_str
        set_progress(90, "Datos actualizados")
    else:
        # Full fetch
        set_progress(2, "Cargando productos...")
        products = fetch_products()
        set_progress(8, "Procesando productos...")
        variant_map, product_names = build_variant_map(products)
        set_progress(10, "Cargando historial completo de pedidos...")
        all_orders = fetch_orders(days=None, progress_range=(10, 90), label="Historial")
        set_progress(92, "Procesando datos...")
        _raw_cache["products"] = products
        _raw_cache["all_orders"] = all_orders
        _raw_cache["variant_map"] = variant_map
        _raw_cache["product_names"] = product_names
        _raw_cache["last_updated"] = now_str

def compute_summary(days=None, date_from=None, date_to=None):
    """Compute summary from _raw_cache for any period — no API calls."""
    products = _raw_cache["products"]
    all_orders = _raw_cache["all_orders"]
    variant_map = _raw_cache["variant_map"]
    product_names = _raw_cache["product_names"]

    if date_from and date_to:
        d1 = datetime.strptime(date_from, "%Y-%m-%d")
        d2 = datetime.strptime(date_to, "%Y-%m-%d")
        days = max((d2 - d1).days, 1)
        cutoff = d1
        cutoff_str = date_from
    else:
        cutoff = datetime.now() - timedelta(days=days)
        cutoff_str = cutoff.strftime("%Y-%m-%d")

    # Filter period orders from full history
    period_orders = [o for o in all_orders if o.get("created_at", "")[:10] >= cutoff_str]

    period_variants_sold = get_variants_with_sales(period_orders)
    all_variants_sold = get_variants_with_sales(all_orders)

    all_dates_map = defaultdict(list)
    for order in all_orders:
        order_date = order.get("created_at", "")[:10]
        if not order_date: continue
        for item in order.get("products", []):
            pid = item.get("product_id")
            vid = item.get("variant_id")
            if pid and vid and int(item.get("quantity", 0)) > 0:
                key = (pid, vid)
                if order_date not in all_dates_map[key]:
                    all_dates_map[key].append(order_date)

    sales = defaultdict(lambda: {"units": 0, "revenue": 0.0, "product_name": "", "variant_name": "", "sale_dates": []})
    total_revenue = 0.0
    total_shipping_cost = 0.0
    shipping_orders = 0
    shipping_by_month = {}

    for order in period_orders:
        total_revenue += float(order.get("total", 0) or 0)
        cost_owner = float(order.get("shipping_cost_owner", 0) or 0)
        cost_customer = float(order.get("shipping_cost_customer", 0) or 0)
        pickup_type = order.get("shipping_pickup_type", "")
        if cost_owner > 0 and cost_customer == 0 and pickup_type == "ship":
            total_shipping_cost += cost_owner
            shipping_orders += 1
        for item in order.get("products", []):
            pid = item.get("product_id")
            vid = item.get("variant_id")
            qty = int(item.get("quantity", 1))
            price = float(item.get("price", 0) or 0)
            key = (pid, vid)
            sales[key]["units"] += qty
            sales[key]["revenue"] += price * qty
            order_date = order.get("created_at", "")[:10]
            if order_date and order_date not in sales[key]["sale_dates"]:
                sales[key]["sale_dates"].append(order_date)
            if vid and vid in variant_map:
                sales[key]["product_name"] = variant_map[vid]["product_name"]
                sales[key]["variant_name"] = variant_map[vid]["variant_name"]
            else:
                sales[key]["product_name"] = product_names.get(pid, get_name(item.get("name", "")))
                sales[key]["variant_name"] = get_name(item.get("variant", ""))

    for order in all_orders:
        cost_owner = float(order.get("shipping_cost_owner", 0) or 0)
        cost_customer = float(order.get("shipping_cost_customer", 0) or 0)
        pickup_type = order.get("shipping_pickup_type", "")
        if pickup_type == "ship":
            month_key = order.get("created_at", "")[:7]
            if month_key:
                if month_key not in shipping_by_month:
                    shipping_by_month[month_key] = {"costo_tienda": 0.0, "costo_cliente": 0.0, "orders": 0}
                # Lo que asume la tienda: cuando el cliente paga $0
                if cost_owner > 0 and cost_customer == 0:
                    shipping_by_month[month_key]["costo_tienda"] += cost_owner
                    shipping_by_month[month_key]["orders"] += 1
                elif cost_customer > 0:
                    shipping_by_month[month_key]["costo_cliente"] += cost_customer

    stagnant = []
    for vid, v in variant_map.items():
        if v["stock"] <= 0: continue
        if not v["variant_name"] or v["variant_name"] == "(sin variante)": continue
        if vid in all_variants_sold: continue
        d = v["days_in_catalog"]
        tipo = "critico" if d >= 180 else "observacion" if d >= 60 else "nuevo"
        stagnant.append({
            "product": v["product_name"], "variant": v["variant_name"],
            "stock": v["stock"], "days_in_catalog": d, "tipo": tipo,
            "image": v.get("image", ""), "price": v["price"], "promo_price": v["promo_price"]
        })
    tipo_order = {"critico": 0, "observacion": 1, "nuevo": 2}
    stagnant.sort(key=lambda x: (tipo_order[x["tipo"]], -x["days_in_catalog"]))

    by_product = defaultdict(lambda: {"name": "", "units": 0, "revenue": 0.0, "variants": []})
    for (pid, vid), s in sales.items():
        rate = round(s["units"] / days, 2)
        stock = variant_map.get(vid, {}).get("stock", 0) if vid else 0
        dias_stock = round(stock / rate) if rate > 0 else None
        by_product[pid]["name"] = s["product_name"]
        by_product[pid]["units"] += s["units"]
        by_product[pid]["revenue"] += s["revenue"]
        if s["variant_name"]:
            by_product[pid]["variants"].append({
                "variant_name": s["variant_name"],
                "units": s["units"],
                "revenue": round(s["revenue"], 2),
                "rate": rate,
                "stock": stock,
                "dias_stock": dias_stock,
                "image": variant_map.get(vid, {}).get("image", "") if vid else "",
                "has_promo": variant_map.get(vid, {}).get("promo_price", 0) > 0 if vid else False,
                "sale_dates": sorted(s["sale_dates"], reverse=True),
                "historical_rate": _calc_historical_rate(all_dates_map.get((pid, vid), []))
            })

    out = []
    for pid, ps in by_product.items():
        ps["variants"].sort(key=lambda x: x["units"], reverse=True)
        has_promo = any(v.get("has_promo", False) for v in ps["variants"])
        out.append({"id": pid, "name": ps["name"], "units": ps["units"], "revenue": round(ps["revenue"], 2), "variants": ps["variants"], "has_promo": has_promo})
    out.sort(key=lambda x: x["units"], reverse=True)

    total_orders = len(period_orders)
    return {
        "days": days,
        "total_orders": total_orders,
        "total_units": sum(p["units"] for p in out),
        "total_revenue": round(total_revenue, 2),
        "ticket_promedio": round(total_revenue / total_orders, 2) if total_orders else 0,
        "products": out,
        "stagnant": stagnant,
        "shipping_cost": round(total_shipping_cost, 2),
        "shipping_orders": shipping_orders,
        "shipping_by_month": {k: {"costo_tienda": round(v["costo_tienda"], 2), "costo_cliente": round(v["costo_cliente"], 2), "orders": v["orders"]} for k, v in sorted(shipping_by_month.items(), reverse=True)}
    }

def build_summary(days=None, date_from=None, date_to=None):
    """Fetch raw data if needed, then compute summary."""
    if _raw_cache["all_orders"] is None:
        fetch_raw_data()
    set_progress(92, "Calculando resumen...")
    result = compute_summary(days=days, date_from=date_from, date_to=date_to)
    set_progress(100, "Listo")
    return result

def _parse_demand_rows(rows, dias_filtro=None):
    """Convert iterable of CSV row dicts to demand dict, deduplicating by email per variant.
    dias_filtro: if set, only consider rows from the last N days (based on createdDate).
    """
    cutoff = None
    if dias_filtro:
        cutoff = datetime.now() - timedelta(days=dias_filtro)

    # raw[key] = {email -> status}
    raw = {}
    for row in rows:
        # Date filter
        if cutoff:
            fecha_str = row.get("createdDate", "")
            if fecha_str:
                try:
                    fecha = datetime.fromisoformat(fecha_str.replace("Z", "+00:00")).replace(tzinfo=None)
                    if fecha < cutoff:
                        continue
                except:
                    pass

        key = (row.get("productName","").strip(), row.get("productVariantName","").strip())
        email = row.get("email","").strip().lower()
        status = row.get("status","").strip()
        if not email:
            continue
        if key not in raw:
            raw[key] = {}
        # Keep "unsent" over "sent" if same email appears twice
        if email not in raw[key] or status == "unsent":
            raw[key][email] = status

    demand = {}
    for key, emails in raw.items():
        demand[key] = {
            "total": len(emails),
            "pendientes": sum(1 for s in emails.values() if s == "unsent")
        }
    return demand

def load_demand_csv(filepath):
    """Load notifications CSV and return dict: (productName, variantName) -> {total, pendientes}"""
    if not filepath or not os.path.exists(filepath):
        return {}
    try:
        with open(filepath, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f, delimiter=";")
            return _parse_demand_rows(reader)
    except Exception as e:
        print(f"  [CSV] Error leyendo demanda: {e}")
        return {}

DEMAND_CSV_PATH = os.environ.get("DEMAND_CSV", "")

def build_export_xlsx(summary_data, demand):
    """Build Excel in memory, return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock y Demanda"

    brown_mid = "7A5C50"
    brown_dark = "3A2A24"
    brown_light = "E8D8CF"

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=brown_mid)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    subheader_font = Font(name="Arial", bold=True, color=brown_dark, size=9)
    subheader_fill = PatternFill("solid", fgColor=brown_light)
    normal_font = Font(name="Arial", size=9, color=brown_dark)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="D0C0B8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.merge_cells("A1:I1")
    ws["A1"] = "San Pretta · Stock y Demanda · " + datetime.now().strftime("%d/%m/%Y")
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color=brown_dark)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="F5EDE6")
    ws.row_dimensions[1].height = 28

    # Headers row 3
    headers = ["Artículo", "Talle", "Stock Restante", "Días sin Ventas",
               "Días en Catálogo", "Notif. Totales", "Notif. Pendientes", "Índice Demanda", "Alerta"]
    col_widths = [40, 8, 13, 13, 14, 13, 15, 13, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[3].height = 32

    row = 4
    for product in summary_data.get("products", []):
        pname = product["name"]
        variants = product.get("variants", [])
        if not variants:
            continue

        # Product sub-header
        for c in range(1, 10):
            cell = ws.cell(row=row, column=c)
            cell.fill = subheader_fill
            cell.border = border
        ws.cell(row=row, column=1, value=pname).font = subheader_font
        ws.cell(row=row, column=1).alignment = left_align
        ws.merge_cells(f"A{row}:I{row}")
        ws.row_dimensions[row].height = 18
        row += 1

        for v in variants:
            vname = v.get("variant_name", "")
            stock = v.get("stock", 0)

            sale_dates = v.get("sale_dates", [])
            if sale_dates:
                try:
                    last = datetime.strptime(sorted(sale_dates)[-1], "%Y-%m-%d")
                    dias_sin_ventas = (datetime.now() - last).days
                except:
                    dias_sin_ventas = None
            else:
                dias_sin_ventas = None

            dias_catalogo = v.get("days_in_catalog", None)

            d_key = (pname, vname)
            dem = demand.get(d_key, {})
            notif_total = dem.get("total", 0)
            notif_pend = dem.get("pendientes", 0)

            if stock == 0 and notif_pend > 0:
                alerta = "Sin stock c/demanda"
                row_fill = PatternFill("solid", fgColor="FAEAEA")
            elif notif_total > 10:
                alerta = "Alta demanda"
                row_fill = PatternFill("solid", fgColor="FFF3E0")
            elif dias_sin_ventas is not None and dias_sin_ventas > 60 and stock > 0:
                alerta = "Sin movimiento"
                row_fill = PatternFill("solid", fgColor="F5F5F5")
            elif notif_total > 0:
                alerta = "Con interés"
                row_fill = None
            else:
                alerta = ""
                row_fill = None

            values = [pname, vname, stock, dias_sin_ventas, dias_catalogo,
                      notif_total or None, notif_pend or None, notif_total or None, alerta]

            for c_idx, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=c_idx, value=val)
                cell.font = normal_font
                cell.border = border
                cell.alignment = center_align if c_idx > 1 else left_align
                if row_fill:
                    cell.fill = row_fill
            ws.row_dimensions[row].height = 16
            row += 1

    ws.freeze_panes = "A4"

    # Sheet 2: Demand ranking
    ws2 = wb.create_sheet("Ranking Demanda")
    h2 = ["Artículo", "Talle", "Notif. Totales", "Notif. Pendientes"]
    w2 = [40, 8, 14, 16]
    for col_idx, (h, w) in enumerate(zip(h2, w2), 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws2.column_dimensions[get_column_letter(col_idx)].width = w

    sorted_demand = sorted(demand.items(), key=lambda x: x[1]["pendientes"], reverse=True)
    for r2, ((pn, vn), dv) in enumerate(sorted_demand, 2):
        if dv["total"] == 0:
            continue
        fill2 = PatternFill("solid", fgColor="FFF3E0") if dv["pendientes"] > 5 else None
        for c_idx, val in enumerate([pn, vn, dv["total"], dv["pendientes"]], 1):
            cell = ws2.cell(row=r2, column=c_idx, value=val)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center_align if c_idx > 1 else left_align
            if fill2:
                cell.fill = fill2

    # Sheet 3: Stock quieto (estancado)
    ws3 = wb.create_sheet("Stock Quieto")
    h3 = ["Artículo", "Talle", "Stock", "Precio", "Precio Promo", "Días en Catálogo", "Estado"]
    w3 = [40, 8, 8, 12, 13, 16, 14]
    for col_idx, (h, w) in enumerate(zip(h3, w3), 1):
        cell = ws3.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws3.column_dimensions[get_column_letter(col_idx)].width = w

    tipo_label = {"critico": "Crítico", "observacion": "Observación", "nuevo": "Nuevo"}
    tipo_fill = {
        "critico": PatternFill("solid", fgColor="FAEAEA"),
        "observacion": PatternFill("solid", fgColor="FFF3E0"),
        "nuevo": None,
    }

    price_fmt = '#,##0'

    stagnant = summary_data.get("stagnant", [])
    for r3, item in enumerate(stagnant, 2):
        tipo = item.get("tipo", "nuevo")
        row_fill3 = tipo_fill.get(tipo)
        price = item.get("price", None) or None
        promo = item.get("promo_price", None) or None
        vals = [
            item.get("product", ""),
            item.get("variant", ""),
            item.get("stock", 0),
            price,
            promo,
            item.get("days_in_catalog", None),
            tipo_label.get(tipo, tipo),
        ]
        for c_idx, val in enumerate(vals, 1):
            cell = ws3.cell(row=r3, column=c_idx, value=val)
            cell.font = normal_font
            cell.border = border
            cell.alignment = center_align if c_idx > 1 else left_align
            if row_fill3:
                cell.fill = row_fill3
            if c_idx in (4, 5) and val:
                cell.number_format = price_fmt

    ws3.freeze_panes = "A2"


    # Sheet 4: Reposición Sugerida
    ws4 = wb.create_sheet("Reposición Sugerida")
    horizonte = summary_data.get("horizonte_reposicion", 60)
    dias_filtro_csv = summary_data.get("dias_filtro_csv", None)
    filtro_label = f"últimos {dias_filtro_csv}d" if dias_filtro_csv else "todo el CSV"

    h4 = ["Artículo", "Talle", "¿Reponer?", "Unidades a Reponer",
          "Stock Actual", f"Ventas Proy. ({horizonte}d)", f"Notif. Pendientes ({filtro_label})", "Prioridad", "Razón"]
    w4 = [40, 8, 10, 18, 12, 18, 24, 12, 55]
    for col_idx, (h, w) in enumerate(zip(h4, w4), 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws4.column_dimensions[get_column_letter(col_idx)].width = w
    ws4.row_dimensions[1].height = 32

    reposicion_rows = []
    for product in summary_data.get("products", []):
        pname = product["name"]
        for v in product.get("variants", []):
            vname = v.get("variant_name", "")
            stock = v.get("stock", 0)
            hist_rate = v.get("historical_rate", 0) or v.get("rate", 0) or 0
            ventas_proyectadas = round(hist_rate * horizonte)
            ventas_por_semana = round(hist_rate * 7, 1)

            d_key = (pname, vname)
            dem = demand.get(d_key, {})
            demanda_reprimida = dem.get("pendientes", 0)
            notif_total = dem.get("total", 0)

            necesidad = max(ventas_proyectadas, demanda_reprimida)
            reposicion = max(0, necesidad - stock)
            reponer = reposicion > 0

            # Prioridad
            if stock == 0 and demanda_reprimida >= 5:
                prioridad = "URGENTE"; prio_order = 0
            elif stock == 0 and reponer:
                prioridad = "Alta"; prio_order = 1
            elif reponer and demanda_reprimida > 0:
                prioridad = "Media"; prio_order = 2
            elif reponer:
                prioridad = "Baja"; prio_order = 3
            else:
                prioridad = "OK"; prio_order = 4

            # Build reason text
            razones = []
            if demanda_reprimida > 0:
                razones.append(f"{demanda_reprimida} persona{'s' if demanda_reprimida > 1 else ''} esperando stock")
            if notif_total > demanda_reprimida and notif_total > 0:
                razones.append(f"{notif_total} notificaciones en total")
            if ventas_por_semana > 0:
                razones.append(f"vende ~{ventas_por_semana}/semana")
            if stock == 0:
                razones.append("sin stock actualmente")
            elif stock > 0 and reposicion > 0:
                razones.append(f"stock actual cubre solo parte de la demanda")
            razon = " · ".join(razones) if razones else "—"

            if not reponer and demanda_reprimida == 0 and ventas_proyectadas == 0:
                continue

            reposicion_rows.append({
                "pname": pname, "vname": vname,
                "reponer": "Sí" if reponer else "No",
                "reposicion": reposicion if reponer else 0,
                "stock": stock, "ventas_proy": ventas_proyectadas,
                "demanda_rep": demanda_reprimida or None,
                "prioridad": prioridad, "prio_order": prio_order,
                "razon": razon
            })

    reposicion_rows.sort(key=lambda x: (x["prio_order"], -x["reposicion"]))

    prio_fills = {
        "URGENTE": PatternFill("solid", fgColor="FAEAEA"),
        "Alta":    PatternFill("solid", fgColor="FDEBD0"),
        "Media":   PatternFill("solid", fgColor="FFF9E6"),
        "Baja":    PatternFill("solid", fgColor="F5F5F5"),
        "OK":      None,
    }
    prio_fonts = {
        "URGENTE": Font(name="Arial", size=9, bold=True, color="C0392B"),
        "Alta":    Font(name="Arial", size=9, bold=True, color="E67E22"),
        "Media":   Font(name="Arial", size=9, color=brown_dark),
        "Baja":    Font(name="Arial", size=9, color=brown_dark),
        "OK":      Font(name="Arial", size=9, color="888888"),
    }

    for r4, row in enumerate(reposicion_rows, 2):
        p = row["prioridad"]
        row_fill4 = prio_fills.get(p)
        vals = [row["pname"], row["vname"], row["reponer"], row["reposicion"] or None,
                row["stock"], row["ventas_proy"] or None, row["demanda_rep"],
                p, row["razon"]]
        for c_idx, val in enumerate(vals, 1):
            cell = ws4.cell(row=r4, column=c_idx, value=val)
            is_prio_col = c_idx == 8
            cell.font = prio_fonts.get(p, normal_font) if is_prio_col else normal_font
            cell.border = border
            cell.alignment = left_align if c_idx in (1, 9) else center_align
            if row_fill4:
                cell.fill = row_fill4
        ws4.row_dimensions[r4].height = 18

    ws4.freeze_panes = "A2"


    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


_cache = {}

def _init_cache_from_drive():
    """Load cache from Gist on startup, then do incremental update to get fresh data."""
    data = drive_load_cache()
    if data:
        days = data.get("days", 90)
        key = f"s{days}"
        _cache[key] = data
        print(f"  [Cache] Cargado desde Gist: key={key}, last_updated={_raw_cache['last_updated']}")

        # Si tenemos órdenes en caché, hacer fetch incremental para actualizarlas
        if _raw_cache["all_orders"] is not None and _raw_cache["last_updated"]:
            print(f"  [Cache] Actualizando órdenes desde {_raw_cache['last_updated']}...")
            try:
                fetch_raw_data(incremental=True)
                _cache.clear()  # Invalidar resumen viejo
                summary = compute_summary(days=90)
                _cache["s90"] = summary
                to_save = dict(summary)
                to_save["_last_updated"] = _raw_cache.get("last_updated", "")
                drive_save_cache(to_save)
                print("  [Cache] Actualización incremental completada")
            except Exception as e:
                print(f"  [Cache] Error en actualización incremental: {e}")
    else:
        print("  [Cache] Sin datos previos en Gist, primera carga completa")

class Handler(http.server.BaseHTTPRequestHandler):
    def log_message(self, *a): pass
    def send_cors(self):
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, OPTIONS")
    def do_HEAD(self):
        # UptimeRobot y otros monitores usan HEAD — responder 200 OK
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_cors()
        self.end_headers()

    def do_OPTIONS(self):
        self.send_response(200); self.send_cors(); self.end_headers()

    def do_POST(self):
        if self.path == "/login":
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length).decode()
            params = {}
            for part in body.split("&"):
                if "=" in part:
                    k, v = part.split("=", 1)
                    params[k] = v.replace("+", " ").replace("%21", "!")
            username = params.get("username", "")
            password = params.get("password", "")
            pw_hash = hashlib.sha256(password.encode()).hexdigest()
            if username == USERNAME and pw_hash == PASSWORD_HASH:
                token = secrets.token_hex(32)
                SESSIONS.add(token)
                self.send_response(302)
                self.send_header("Location", "/")
                self.send_header("Set-Cookie", f"sp_session={token}; Path=/; HttpOnly; SameSite=Lax")
                self.end_headers()
            else:
                self.send_response(200)
                self.send_header("Content-Type", "text/html")
                self.end_headers()
                html = LOGIN_HTML.replace("{error}", '<div class="error">Usuario o contraseña incorrectos</div>')
                self.wfile.write(html.encode())
        elif self.path == "/export_upload":
            if not check_session(self):
                self.send_response(403); self.end_headers(); return
            self.serve_export_upload()
        else:
            self.send_response(404); self.end_headers()
    def do_GET(self):
        # Public: login page
        if self.path == "/login":
            self.send_response(200)
            self.send_header("Content-Type", "text/html")
            self.end_headers()
            self.wfile.write(LOGIN_HTML.replace("{error}", "").encode())
            return
        # Public: keepalive - mantiene el caché caliente
        if self.path == "/keepalive":
            status = {"ok": True, "cache_keys": list(_cache.keys()), "has_raw": _raw_cache["all_orders"] is not None}
            if _raw_cache["all_orders"] is None:
                # Trigger fetch in background so UptimeRobot doesn't timeout
                threading.Thread(target=self._keepalive_fetch, daemon=True).start()
                status["msg"] = "fetch iniciado en background"
            else:
                status["msg"] = f"{len(_raw_cache['all_orders'])} ordenes en memoria"
            data = json.dumps(status).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_cors(); self.end_headers()
            self.wfile.write(data)
            return

        # Public: diagnostico
        if self.path == "/diagnostico":
            result = {}
            result["GITHUB_TOKEN_len"] = len(GITHUB_TOKEN)
            result["cache_keys"] = list(_cache.keys())
            try:
                gist_id = _find_gist_id()
                result["gist_id"] = gist_id or "no encontrado"
                if gist_id:
                    result["gist_url"] = f"https://gist.github.com/{gist_id}"
            except Exception as e:
                result["gist_error"] = str(e)
            data = json.dumps(result, indent=2, ensure_ascii=False).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_cors(); self.end_headers()
            self.wfile.write(data)
            return
        # Protected: everything else
        if not check_session(self):
            self.send_response(302)
            self.send_header("Location", "/login")
            self.end_headers()
            return
        if self.path == "/": self.serve_file("dashboard.html", "text/html")
        elif self.path == "/progress":
            data = json.dumps(_progress).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_cors()
            self.end_headers()
            self.wfile.write(data)
        elif self.path.startswith("/summary"): self.serve_summary()
        elif self.path == "/invalidate":
            _cache.clear()
            self.send_response(200); self.send_cors(); self.end_headers()
            self.wfile.write(b'{"ok":true}')
        elif self.path.startswith("/export"):
            self.serve_export()
        elif self.path == "/diagnostico":
            result = {}
            # Check env vars
            creds_json = os.environ.get("GOOGLE_CREDENTIALS_JSON", "")
            result["GOOGLE_CREDENTIALS_JSON_len"] = len(creds_json)
            result["GDRIVE_FOLDER_ID"] = os.environ.get("GDRIVE_FOLDER_ID", "no configurado")
            result["GDRIVE_OK"] = GDRIVE_OK
            # Try to connect
            try:
                svc = _get_drive_service()
                result["drive_service"] = "OK" if svc else "None"
                if svc:
                    res = svc.files().list(
                        q=f"'{GDRIVE_FOLDER_ID}' in parents and trashed=false",
                        fields="files(id,name)",
                        pageSize=5
                    ).execute()
                    result["archivos_en_carpeta"] = [f["name"] for f in res.get("files", [])]
                    # Try a test write
                    try:
                        from googleapiclient.http import MediaInMemoryUpload as MIU
                        test_media = MIU(b"test", mimetype="text/plain", resumable=False)
                        tf = svc.files().create(
                            body={"name": "_test_write.txt", "parents": [GDRIVE_FOLDER_ID]},
                            media_body=test_media, fields="id"
                        ).execute()
                        svc.files().delete(fileId=tf["id"]).execute()
                        result["test_escritura"] = "OK"
                    except Exception as e:
                        result["test_escritura_error"] = str(e)
            except Exception as e:
                result["drive_error"] = str(e)
            # Check cache
            result["cache_keys"] = list(_cache.keys())
            data = json.dumps(result, indent=2, ensure_ascii=False).encode()
            self.send_response(200)
            self.send_header("Content-Type", "application/json")
            self.send_cors(); self.end_headers()
            self.wfile.write(data)
        else:
            self.send_response(404); self.end_headers()
    def serve_file(self, filename, content_type):
        fp = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
        try:
            with open(fp, "rb") as f: content = f.read()
            self.send_response(200)
            self.send_header("Content-Type", content_type)
            self.send_cors(); self.end_headers()
            self.wfile.write(content)
        except FileNotFoundError:
            self.send_response(404); self.end_headers()
            self.wfile.write(b"Archivo no encontrado")
    def serve_export_upload(self):
        """Receive multipart POST with CSV + days, return xlsx."""
        if not OPENPYXL_OK:
            self.send_response(500); self.end_headers()
            self.wfile.write(b'openpyxl no instalado'); return
        try:
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length)
            content_type = self.headers.get("Content-Type", "")
            # Parse boundary
            boundary = None
            for part in content_type.split(";"):
                part = part.strip()
                if part.startswith("boundary="):
                    boundary = part[9:].strip().encode()
            demand = {}
            days = 90
            horizonte = 60
            dias_filtro_csv = None
            csv_rows = None
            if boundary:
                parts = body.split(b"--" + boundary)
                for part in parts:
                    if b"Content-Disposition" not in part:
                        continue
                    header_end = part.find(b"\r\n\r\n")
                    if header_end == -1:
                        continue
                    header_raw = part[:header_end].decode(errors="replace")
                    data = part[header_end+4:].rstrip(b"\r\n--")
                    if 'name="csv"' in header_raw:
                        csv_text = data.decode("utf-8-sig", errors="replace")
                        csv_rows = list(csv.DictReader(io.StringIO(csv_text), delimiter=";"))
                    elif 'name="days"' in header_raw:
                        try:
                            days = int(data.decode().strip())
                        except:
                            days = 90
                    elif 'name="horizonte"' in header_raw:
                        try:
                            horizonte = int(data.decode().strip())
                        except:
                            horizonte = 60
                    elif 'name="dias_filtro_csv"' in header_raw:
                        try:
                            v = data.decode().strip()
                            dias_filtro_csv = int(v) if v != "0" else None
                        except:
                            dias_filtro_csv = None
            if csv_rows is not None:
                demand = _parse_demand_rows(csv_rows, dias_filtro=dias_filtro_csv)
            key = f"s{days}"
            if key not in _cache:
                set_progress(0, "Iniciando...")
                _cache[key] = build_summary(days)
                set_progress(100, "Listo")
            summary_with_horizonte = dict(_cache[key])
            summary_with_horizonte["horizonte_reposicion"] = horizonte
            summary_with_horizonte["dias_filtro_csv"] = dias_filtro_csv
            xlsx_bytes = build_export_xlsx(summary_with_horizonte, demand)
            filename = f"sanpretta_stock_{datetime.now().strftime('%Y%m%d')}.xlsx"
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(xlsx_bytes)))
            self.send_cors()
            self.end_headers()
            self.wfile.write(xlsx_bytes)
        except Exception as e:
            print(f"  [export_upload] Error: {e}")
            self.send_response(500); self.end_headers()
            self.wfile.write(str(e).encode())

    def _keepalive_fetch(self):
        """Background fetch triggered by keepalive ping."""
        try:
            if _raw_cache["all_orders"] is None:
                set_progress(0, "Keepalive: cargando datos...")
                fetch_raw_data(incremental=_raw_cache["last_updated"] is not None)
                _cache.clear()
                summary = compute_summary(days=90)
                _cache["s90"] = summary
                to_save = dict(summary)
                to_save["_last_updated"] = _raw_cache.get("last_updated", "")
                drive_save_cache(to_save)
                set_progress(100, "Listo")
                print("  [Keepalive] Datos cargados en background")
        except Exception as e:
            print(f"  [Keepalive] Error: {e}")

    def serve_export(self):
        if not OPENPYXL_OK:
            self.send_response(500); self.end_headers()
            self.wfile.write(b'openpyxl no instalado')
            return
        qs = parse_qs(urlparse(self.path).query)
        days = int(qs.get("days", ["90"])[0])
        key = f"s{days}"
        if key not in _cache:
            set_progress(0, "Iniciando...")
            _cache[key] = build_summary(days)
            set_progress(100, "Listo")
        csv_path = qs.get("csv", [DEMAND_CSV_PATH])[0]
        horizonte = int(qs.get("horizonte", ["60"])[0])
        demand = load_demand_csv(csv_path)
        summary_with_horizonte = dict(_cache[key])
        summary_with_horizonte["horizonte_reposicion"] = horizonte
        xlsx_bytes = build_export_xlsx(summary_with_horizonte, demand)
        filename = f"sanpretta_stock_{datetime.now().strftime('%Y%m%d')}.xlsx"
        self.send_response(200)
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
        self.send_header("Content-Length", str(len(xlsx_bytes)))
        self.send_cors()
        self.end_headers()
        self.wfile.write(xlsx_bytes)

    def serve_summary(self):
        qs = parse_qs(urlparse(self.path).query)
        date_from = qs.get("date_from", [None])[0]
        date_to   = qs.get("date_to",   [None])[0]
        force     = qs.get("force",     ["0"])[0] == "1"

        # Key: custom range or days
        if date_from and date_to:
            key = f"custom_{date_from}_{date_to}"
            days = None
        else:
            days = int(qs.get("days", ["90"])[0])
            key = f"s{days}"

        # Fetch raw data if needed
        if force:
            set_progress(0, "Actualizando todo desde Tiendanube...")
            fetch_raw_data(incremental=False)
            _cache.clear()
        elif _raw_cache["all_orders"] is None:
            set_progress(0, "Iniciando...")
            fetch_raw_data(incremental=_raw_cache["last_updated"] is not None)
            _cache.clear()

        if key not in _cache:
            set_progress(92, f"Calculando resumen...")
            _cache[key] = compute_summary(days=days, date_from=date_from, date_to=date_to)
            set_progress(100, "Listo")
            # Save 90d version to Gist as persistent cache
            save_key = "s90" if "s90" in _cache else key
            to_save = dict(_cache[save_key])
            to_save["_last_updated"] = _raw_cache.get("last_updated", "")
            threading.Thread(target=drive_save_cache, args=(to_save,), daemon=True).start()

        data = json.dumps(_cache[key]).encode()
        self.send_response(200)
        self.send_header("Content-Type", "application/json")
        self.send_cors(); self.end_headers()
        self.wfile.write(data)

if __name__ == "__main__":
    print(f"\n  Dashboard San Pretta")
    print(f"  Cargando caché desde Drive...")
    _init_cache_from_drive()
    server = http.server.ThreadingHTTPServer(("0.0.0.0", PORT), Handler)
    print(f"  Abri http://localhost:{PORT} en tu browser")
    print(f"  Ctrl+C para detener\n")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Servidor detenido.")
